# -*- coding: utf-8 -*-

import xlrd
from openpyxl import load_workbook
from elasticsearch_query import ES
from elasticsearch import Elasticsearch


es_host = '10.84.108.114'
es_port = 9200
index_name = 'test_index'
doc_type = '_doc'


class ExcelDealer(object):
    def __init__(self, path):
        self.path = path
        self.book = None
        self.sheet = None

    def get_excel_book(self):
        self.book = xlrd.open_workbook(self.path)

    def get_sheet_content(self, sheet_name):
        self.sheet = self.book.sheet_by_name(sheet_name)
        experience_error_dict = {}
        rows = self.sheet.nrows
        columns = self.sheet.ncols
        column_list = ['is_jenkins_job_issue', 'is_reinit_issue', 'is_environmental_issue', 'is_test_script_issue',
                       'is_trident_bug', 'component', 'detailed_dc_log', 'key_message']

        for row in range(1, rows):
            data_list = []
            for column in range(1, columns):
                data_list.append(self.sheet.cell_value(row, column))
            experience_error_dict[self.sheet.cell_value(row, 0)] = dict(zip(column_list, data_list))
        return experience_error_dict

    def get_mdt_root_cause_content(self, sheet_name):
        self.sheet = self.book.sheet_by_name(sheet_name)
        experience_error_dict = {}
        rows = self.sheet.nrows
        columns = self.sheet.ncols

        for row in range(1, rows):
            experience_error_dict[self.sheet.cell_value(row, 1)] = self.sheet.cell_value(row, 8)
        return experience_error_dict

    def write_using_openpyxl(self, excel_path, sheet_name, mdt_list):
        wb = load_workbook(excel_path)
        target_sheet = wb.get_sheet_by_name(sheet_name)
        rows = self.sheet.nrows
        row = rows + 1
        for mdt in mdt_list:
            target_sheet.cell(row, 1, mdt)
            row = row + 1
        wb.save(excel_path)

    def get_mdt_list(self):
        es = Elasticsearch([{'host': es_host, 'port': es_port}])
        used_es = ES(es, index_name, doc_type)
        mdt_list = used_es.match_all_query()
        return mdt_list


if __name__ == '__main__':
        excel_path = "./known_issue_list.xlsx"
        dealer = ExcelDealer(excel_path)
        dealer.get_excel_book()
        #sheet_name = 'fe-protocol-ci-warnado-SAN'
        #print dealer.get_sheet_content(sheet_name)
        sheet_name = 'IO Modules Known MDT'
        print dealer.get_mdt_root_cause_content(sheet_name)

        mdt_file_path = 'mdt_list.txt'
        mdt_list = dealer.get_mdt_list()
        print '#####' * 50
        print len(mdt_list)
        print mdt_list
        mdt_list = ['MDT-12345', 'MDT-23456', 'MDT-34567']
        dealer.write_using_openpyxl(excel_path, sheet_name, mdt_list)
